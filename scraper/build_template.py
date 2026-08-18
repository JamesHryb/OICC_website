"""
One-off generator for scraper/templates/super_sixes_template.xlsx.

Not part of the runtime data pipeline — this just builds the starter workbook
that gets uploaded to Google Drive and turned into the scorers' spreadsheet.
Re-run it only if you want to regenerate the template from scratch (requires
`pip install openpyxl`, which is NOT a runtime dependency of the scraper).

Design notes (matches scraper/super_sixes_manual.py's expectations):
  - Fixtures_Results: whichever team you enter as Team1 is the one that
    batted first — there's no separate BattedFirst flag to keep in sync.
  - Batting/Bowling: you only pick MatchNo + Innings (1 or 2) — the Batting
    Team / Bowling Team columns are computed by looking up Fixtures_Results,
    and the Player/Bowler/Fielder dropdowns filter to that resolved team's
    roster automatically. This removes the "pick the right team AND the
    right player, and don't let them drift apart" failure mode entirely.
  - The dependent Player/Bowler/Fielder dropdowns are built from literal
    per-row helper columns (N:W, Z:AI, M:V) rather than sourcing the
    dropdown directly from INDIRECT() — Google Sheets' dropdown-arrow UI
    does not reliably render a populated list when the source is a formula,
    even though the validation logic itself evaluates correctly. Pre-
    resolving each row's roster into literal helper cells and pointing the
    dropdown at *those* sidesteps that limitation.

Usage:
    python build_template.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.formula import ArrayFormula

NAVY = "003F87"
GOLD = "FFD700"
OUT_PATH = Path(__file__).parent / "templates" / "super_sixes_template.xlsx"

# How far down each dropdown/validation/formula extends, so it still applies
# to rows added later in the day rather than only the starter rows below.
MAX_DATA_ROW = 300

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
NOTE_FONT = Font(italic=True, color="6C757D")

TEAM_COUNT = 6
PLAYERS_PER_TEAM = 6
SPARE_PLAYER_ROWS = 4  # blank rows left per team block to add players later without resizing the range
ROSTER_BLOCK_SIZE = PLAYERS_PER_TEAM + SPARE_PLAYER_ROWS  # width the Player/Bowler/Fielder dropdowns must match

MAX_FIXTURE_ROW = 15  # headroom beyond the 11 starter matches, for the MatchNo dropdown lookups


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = f"A{row + 1}"


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w


def add_list_dv(ws, formula1, col_letter, first_row=2, last_row=MAX_DATA_ROW):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")
    return dv


def fill_array_formula(ws, col_letter, formula_template, span_cols, first_row=2, last_row=MAX_DATA_ROW):
    """Write a per-row array formula (e.g. a TRANSPOSE spill) into col_letter
    for every row first_row..last_row. formula_template is an '=...{r}...'
    string where {r} is replaced with the row number. span_cols is how many
    columns wide each row's spill is (1 for a plain formula)."""
    for r in range(first_row, last_row + 1):
        text = formula_template.format(r=r)
        if span_cols > 1:
            end_col = ws.cell(row=r, column=ws[f"{col_letter}1"].column + span_cols - 1).column_letter
            ref = f"{col_letter}{r}:{end_col}{r}"
        else:
            ref = f"{col_letter}{r}"
        ws[f"{col_letter}{r}"] = ArrayFormula(ref, text)


def build():
    wb = Workbook()

    # ── Teams ─────────────────────────────────────────────────
    ws_teams = wb.active
    ws_teams.title = "Teams"
    ws_teams.append(["Team", "Group"])
    style_header(ws_teams, 2)
    groups = ["A", "A", "A", "B", "B", "B"]
    for i in range(1, TEAM_COUNT + 1):
        ws_teams.append([f"Team {i}", groups[(i - 1) % len(groups)]])
    ws_teams.append(["TBC", ""])  # so the Team1/Team2 dropdown accepts the knockout placeholder
    autosize(ws_teams, [20, 10])
    team_range = f"Teams!$A$2:$A${1 + TEAM_COUNT + 1 + 6}"  # a little headroom beyond 6 teams + TBC

    # ── Players ───────────────────────────────────────────────
    # Laid out as one contiguous block of rows per team, so each team's
    # block can be named (Team_1, Team_2, ...) and used to resolve that
    # row's Player/Bowler/Fielder dropdown. Fill in real names ahead of the
    # day; leave the blank rows in each block for subs rather than
    # inserting new rows (inserting/deleting rows can shift a block out
    # from under its name).
    ws_players = wb.create_sheet("Players")
    ws_players.append(["Player", "Team"])
    style_header(ws_players, 2)
    team_player_ranges = {}  # "Team 1" -> "Players!$A$2:$A$11"
    row = 2
    for t in range(1, TEAM_COUNT + 1):
        team_name = f"Team {t}"
        first_row = row
        for p in range(1, PLAYERS_PER_TEAM + 1):
            ws_players.append([f"Player {t}.{p}", team_name])
            row += 1
        for _ in range(SPARE_PLAYER_ROWS):
            ws_players.append(["", team_name])
            row += 1
        team_player_ranges[team_name] = f"Players!$A${first_row}:$A${row - 1}"
    autosize(ws_players, [20, 14])

    # Named range per team, pointing at that team's block above. If a team
    # is renamed in the Teams tab, its named range must be renamed to match
    # (Data > Named ranges in Google Sheets) — the two are matched by name,
    # not by any other link. Team names with punctuation Sheets can't use in
    # a range name (e.g. apostrophes) will need a simpler alias here.
    for team_name, rng in team_player_ranges.items():
        range_name = team_name.replace(" ", "_")
        wb.defined_names[range_name] = DefinedName(range_name, attr_text=rng)

    # ── Fixtures_Results ─────────────────────────────────────
    ws = wb.create_sheet("Fixtures_Results")
    headers = ["MatchNo", "Round", "Team1", "Team2", "Time", "Status",
               "TeamA_Runs", "TeamA_Wickets", "TeamA_Overs",
               "TeamB_Runs", "TeamB_Wickets", "TeamB_Overs", "Notes"]
    ws.append(headers)
    style_header(ws, len(headers))
    rows = [
        # MatchNo, Round, Team1 (batted first), Team2 (batted second), Time, Status, scores..., Notes
        [1, "Group A", "Team 1", "Team 2", "09:30", "Scheduled", "", "", "", "", "", "", ""],
        [2, "Group B", "Team 4", "Team 5", "09:30", "Scheduled", "", "", "", "", "", "", ""],
        [3, "Group A", "Team 1", "Team 3", "10:10", "Scheduled", "", "", "", "", "", "", ""],
        [4, "Group B", "Team 4", "Team 6", "10:10", "Scheduled", "", "", "", "", "", "", ""],
        [5, "Group A", "Team 2", "Team 3", "10:50", "Scheduled", "", "", "", "", "", "", ""],
        [6, "Group B", "Team 5", "Team 6", "10:50", "Scheduled", "", "", "", "", "", "", ""],
        [7, "Semi Final 1", "TBC", "TBC", "13:00", "Scheduled", "", "", "", "", "", "", "Winner Group A v Runner-up Group B"],
        [8, "Semi Final 2", "TBC", "TBC", "13:00", "Scheduled", "", "", "", "", "", "", "Winner Group B v Runner-up Group A"],
        [9, "Wooden Spoon", "TBC", "TBC", "12:20", "Scheduled", "", "", "", "", "", "", "3rd-placed teams in each group"],
        [10, "3rd Place Playoff", "TBC", "TBC", "14:00", "Scheduled", "", "", "", "", "", "", "Losers of the semi-finals"],
        [11, "Final", "TBC", "TBC", "14:40", "Scheduled", "", "", "", "", "", "", "Winners of the semi-finals"],
    ]
    for r in rows:
        ws.append(r)
    autosize(ws, [8, 16, 14, 14, 8, 11, 10, 12, 11, 10, 12, 11, 32])

    add_list_dv(ws, '"Scheduled,Live,Complete,Abandoned,Cancelled"', "F")
    add_list_dv(
        ws,
        '"Group A,Group B,Semi Final 1,Semi Final 2,3rd Place Playoff,Final,Wooden Spoon"',
        "B",
    )
    add_list_dv(ws, team_range, "C")
    add_list_dv(ws, team_range, "D")
    match_range = f"Fixtures_Results!$A$2:$A${MAX_FIXTURE_ROW}"

    # ── Batting ───────────────────────────────────────────────
    # C/D (Batting Team / Bowling Team) are computed from MatchNo + Innings;
    # N:W and Z:AI are hidden-in-spirit helper columns resolving that row's
    # batting-team and bowling-team rosters, used only as dropdown sources.
    ws = wb.create_sheet("Batting")
    headers = ["MatchNo", "Innings", "Batting Team", "Bowling Team", "Player",
               "Runs", "Balls", "Fours", "Sixes", "HowOut", "Bowler", "Fielder"]
    ws.append(headers)
    style_header(ws, len(headers))
    ws.append([1, 1, "", "", "Player 1.1", 24, 15, 3, 1, "not out", "", ""])
    ws.append([1, 1, "", "", "Player 1.2", 18, 12, 2, 0, "ct", "Player 2.1", "Player 2.2"])
    autosize(ws, [8, 9, 16, 16, 16, 8, 8, 8, 8, 10, 14, 14])

    fill_array_formula(
        ws, "C",
        '=iferror(INDEX(Fixtures_Results!$C$2:$D$12,MATCH($A{r},Fixtures_Results!$A$2:$A$12,0),'
        'MATCH("Team"&$B{r},Fixtures_Results!$C$1:$D$1,0)),"")',
        span_cols=1,
    )
    fill_array_formula(
        ws, "D",
        '=iferror(INDEX(Fixtures_Results!$C$2:$D$12,MATCH($A{r},Fixtures_Results!$A$2:$A$12,0),'
        'MATCH("Team"&2*1/$B{r},Fixtures_Results!$C$1:$D$1,0)),"")',
        span_cols=1,
    )
    fill_array_formula(
        ws, "N",
        '=iferror(transpose(INDIRECT(SUBSTITUTE($C{r}," ","_"))),"")',
        span_cols=ROSTER_BLOCK_SIZE,
    )
    fill_array_formula(
        ws, "Z",
        '=iferror(transpose(INDIRECT(SUBSTITUTE($D{r}," ","_"))),"")',
        span_cols=ROSTER_BLOCK_SIZE,
    )

    add_list_dv(ws, match_range, "A")
    add_list_dv(ws, '"1,2"', "B")
    add_list_dv(ws, "Batting!$N2:$" + ws.cell(row=2, column=14 + ROSTER_BLOCK_SIZE - 1).column_letter + "2", "E")
    add_list_dv(ws, '"not out,b,ct,lbw,run out,st,hw"', "J")
    bowler_col_end = ws.cell(row=2, column=26 + ROSTER_BLOCK_SIZE - 1).column_letter  # Z=26
    add_list_dv(ws, f"Batting!$Z2:${bowler_col_end}2", "K")
    add_list_dv(ws, f"Batting!$Z2:${bowler_col_end}2", "L")

    # ── Bowling ───────────────────────────────────────────────
    ws = wb.create_sheet("Bowling")
    headers = ["MatchNo", "Innings", "Bowling Team", "Bowler", "Overs", "Maidens", "Runs", "Wickets"]
    ws.append(headers)
    style_header(ws, len(headers))
    ws.append([1, 1, "", "Player 2.1", "3.0", 0, 24, 2])
    autosize(ws, [8, 9, 16, 16, 8, 9, 8, 9])

    fill_array_formula(
        ws, "C",
        '=iferror(INDEX(Fixtures_Results!$C$2:$D$12,MATCH($A{r},Fixtures_Results!$A$2:$A$12,0),'
        'MATCH("Team"&2*1/$B{r},Fixtures_Results!$C$1:$D$1,0)),"")',
        span_cols=1,
    )
    fill_array_formula(
        ws, "M",
        '=iferror(transpose(INDIRECT(SUBSTITUTE($C{r}," ","_"))),"")',
        span_cols=ROSTER_BLOCK_SIZE,
    )

    add_list_dv(ws, match_range, "A")
    add_list_dv(ws, '"1,2"', "B")
    bowler_col_end2 = ws.cell(row=2, column=13 + ROSTER_BLOCK_SIZE - 1).column_letter  # M=13
    add_list_dv(ws, f"Bowling!$M2:${bowler_col_end2}2", "D")

    # ── Instructions (created last, then moved to the first tab position) ──
    ws = wb.create_sheet("Instructions", 0)
    ws["A1"] = "Imperial Super Sixes — Scorer's Spreadsheet"
    ws["A1"].font = TITLE_FONT
    lines = [
        "",
        "How this works:",
        "1. Fill in the Teams tab once, before the day (team names + which group they're in).",
        "2. Fill in the Players tab once, before the day (one row per player + their team).",
        "3. Fill in Fixtures_Results: the running order is already sketched out below — edit team",
        "   names for the knockout rounds once group results are known. Whichever team you put in",
        "   the Team1 column is the one that batted first — there's no separate field for that.",
        "4. After each match, update its row in Fixtures_Results (Status + both teams' scores),",
        "   then on Batting/Bowling just pick the MatchNo and Innings (1 or 2) for each entry — the",
        "   Team and Player/Bowler/Fielder columns become dropdowns automatically, already filtered",
        "   to the right teams, once step 3 is filled in for that match.",
        "5. The website re-reads this spreadsheet automatically every few minutes. To force an",
        "   immediate refresh, use the 'Run workflow' button on the GitHub Actions tab.",
        "",
        "Column notes:",
        "  Status (Fixtures_Results): Scheduled / Live / Complete / Abandoned / Cancelled",
        "  Round: Group A / Group B / Semi Final 1 / Semi Final 2 / 3rd Place Playoff / Final /",
        "    Wooden Spoon — must match exactly (including capitalisation) for the bracket to work.",
        "  Innings (Batting/Bowling): 1 or 2 — which of the match's two innings this row belongs to.",
        "  HowOut (Batting): not out / b / ct / lbw / run out / st / hw",
        "  Overs (Bowling): cricket notation, e.g. 3.4 = 3 overs and 4 balls (not 3.4 overs decimal).",
        "",
        "Matches are 5 overs a side — the website uses this to compute Net Run Rate correctly for",
        "teams bowled out early. If the format changes, say so and the site config can be updated.",
        "",
        "Keep team and player names byte-for-byte identical across tabs — use the dropdowns rather",
        "than retyping, or the group table and stats will split someone into two separate entries.",
        "",
        "Full setup guide: scraper/SUPER_SIXES_GUIDE.md in the website repository.",
    ]
    for i, line in enumerate(lines, start=2):
        cell = ws.cell(row=i, column=1, value=line)
        if line.strip().endswith(":"):
            cell.font = Font(bold=True, color=NAVY)
    ws.column_dimensions["A"].width = 100

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Saved: {OUT_PATH}")


if __name__ == "__main__":
    build()
